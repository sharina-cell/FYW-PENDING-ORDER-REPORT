"""
FYW Pending Order Report Generator
===================================
Core logic for generating the daily FYW pending order Excel report.
Used by both the CLI script and the Streamlit app.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import io


# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
PENDING_STATUSES = ['ACCEPTED/PICKED', 'READY TO SHIP', 'NEW', 'CANCEL REQUESTED']

BRAND_COLORS  = {'MELISSA': 'D6E4F7', 'IPANEMA': 'D6F0E0', 'CSPACE': 'FFF2CC'}
HEADER_BG     = '1F4E79'
SUBHEADER_BG  = {'MELISSA': '2E75B6', 'IPANEMA': '375623', 'CSPACE': 'BF8F00'}
STATUS_COLORS = {
    'ACCEPTED/PICKED':  'E2EFDA',
    'READY TO SHIP':    'FFF2CC',
    'NEW':              'DDEEFF',
    'CANCEL REQUESTED': 'FCE4D6',
}
NICK_BG = {
    'shopee-Melissa MY':        '2E75B6',
    'shopee-Ipanema MY':        '375623',
    'shopee-CSpace MY':         'BF8F00',
    'tiktok-tiktok-CspaceMY':   '444444',
    'tiktok-MelissaMY':         '444444',
    'tiktok-tiktok-IpanemaMY':  '444444',
    'zalora-Melissa-MY':        'C00000',
    'zalora-Ipanema-MY':        'C00000',
    'lazada-Melissa MY':        '0F146D',
    'lazada-Ipanema MY':        '0F146D',
}
BRAND_ORDER = {'MELISSA': 0, 'IPANEMA': 1, 'CSPACE': 2, 'OTHER': 3}


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def get_brand(nickname):
    if pd.isna(nickname): return 'OTHER'
    n = str(nickname).lower()
    if 'melissa' in n: return 'MELISSA'
    if 'ipanema' in n: return 'IPANEMA'
    if 'cspace'  in n: return 'CSPACE'
    return 'OTHER'


def get_channel(nickname):
    if pd.isna(nickname): return ''
    n = str(nickname).lower()
    if 'shopee' in n: return 'Shopee'
    if 'tiktok' in n: return 'TikTok'
    if 'lazada' in n: return 'Lazada'
    if 'zalora' in n: return 'Zalora'
    return str(nickname)


def classify_col(col_name, today, yesterday):
    """Classify a date column as past / yesterday / today / future / grand_total."""
    if col_name == 'Grand Total': return 'grand_total'
    try:
        col_date = datetime.strptime(col_name, '%d/%m/%Y').date()
        if col_date < yesterday:  return 'past'
        if col_date == yesterday: return 'yesterday'
        if col_date == today:     return 'today'
        return 'future'
    except:
        return 'future'


def col_header_bg(col_name, today, yesterday):
    c = classify_col(col_name, today, yesterday)
    return {'past':'C00000','yesterday':'C55A00','today':'375623','grand_total':'2E75B6'}.get(c,'1F4E79')


def col_data_bg(col_name, today, yesterday):
    c = classify_col(col_name, today, yesterday)
    return {'past':'FFCCCC','yesterday':'FCE4D6','today':'C6EFCE','grand_total':'DDEEFF'}.get(c,'F2F2F2')


def col_data_font_color(col_name, today, yesterday):
    c = classify_col(col_name, today, yesterday)
    return {'past':'9C0006','yesterday':'833C00','today':'006100'}.get(c,'000000')


def col_grand_bg(col_name, today, yesterday):
    c = classify_col(col_name, today, yesterday)
    return {'past':'C00000','yesterday':'C55A00','today':'70AD47','grand_total':'2E75B6'}.get(c,'1F4E79')


def hdr_style(cell, bg=HEADER_BG, fg='FFFFFF', size=9):
    thin = Side(style='thin', color='CCCCCC')
    cell.font      = Font(name='Arial', bold=True, color=fg, size=size)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)


# ─────────────────────────────────────────────
# MAIN REPORT BUILDER
# ─────────────────────────────────────────────
def build_report(all_orders_file, shopee_files: dict, report_date: datetime = None) -> tuple[bytes, dict]:
    """
    Build the FYW Pending Order Report.

    Parameters
    ----------
    all_orders_file : str | file-like
        Path or file object for the FYW ALL CSV export.
    shopee_files : dict
        Dict of brand → file path or file-like object.
        e.g. {'melissa': 'MELISSA_SHOPEE_ORDER.xlsx', 'ipanema': ..., 'cspace': ...}
        Missing brands are skipped gracefully.
    report_date : datetime, optional
        Override today's date (useful for testing). Defaults to datetime.today().

    Returns
    -------
    bytes
        The Excel workbook as bytes (ready for download or saving).
    dict
        Summary stats: total_items, total_value, nicknames, sla_dates, pivot_df.
    """
    today     = (report_date or datetime.today()).date()
    yesterday = today - timedelta(days=1)
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Load FYW CSV ──────────────────────────────────────────────────────────
    all_orders = pd.read_csv(all_orders_file)
    all_orders['order_id_clean']  = all_orders['order_id'].str.strip()
    all_orders['tracking_clean']  = all_orders['tracking_number'].str.strip()

    pending = all_orders[all_orders['order_status'].isin(PENDING_STATUSES)].copy()
    pending['order_id_clean']  = pending['order_id'].str.strip()
    pending['tracking_clean']  = pending['tracking_number'].str.strip()
    pending['Nickname']        = pending['nickname'].str.strip()
    pending['Brand']           = pending['nickname'].apply(get_brand)
    pending['Channel']         = pending['nickname'].apply(get_channel)

    # ── Load Shopee files ─────────────────────────────────────────────────────
    mp_sla_map      = {}
    shopee_tracking = {}
    load_warnings   = []

    for brand_key, filepath in shopee_files.items():
        if filepath is None:
            continue
        try:
            df = pd.read_excel(filepath)
            for _, row in df.iterrows():
                oid = str(row['Order ID']).strip()
                est = row.get('Estimated Ship Out Date', None)
                if pd.notna(est) and str(est).strip() not in ['', 'nan']:
                    mp_sla_map[oid] = str(est).strip()
                trk = str(row['Tracking Number*']).strip() if pd.notna(row['Tracking Number*']) else ''
                if trk and trk != 'nan':
                    shopee_tracking[oid] = trk
        except Exception as e:
            load_warnings.append(f'⚠️  Could not load {brand_key} Shopee file: {e}')

    # ── Build report rows ─────────────────────────────────────────────────────
    report_rows = []
    for _, row in pending.iterrows():
        oid     = row['order_id_clean']
        channel = row['Channel']
        tracking = row['tracking_clean']
        if not tracking or tracking == 'nan':
            tracking = shopee_tracking.get(oid, '')

        try:
            ord_dt   = pd.to_datetime(row['ordered_date'], dayfirst=True)
            ord_date = ord_dt.strftime('%d/%m/%Y')
        except:
            ord_dt   = None
            ord_date = str(row['ordered_date'])

        try:
            sla_date = pd.to_datetime(row['order_sla'], dayfirst=True).strftime('%d/%m/%Y')
        except:
            sla_date = str(row['order_sla']) if pd.notna(row['order_sla']) else ''

        # MP SLA logic
        if channel == 'Shopee':
            raw = mp_sla_map.get(oid, '')
            try:
                mp_sla = pd.to_datetime(raw).strftime('%d/%m/%Y %H:%M') if raw else ''
            except Exception:
                mp_sla = raw
        elif channel in ('TikTok', 'Lazada', 'Zalora'):
            # derive MP SLA from order date (fallback parse if initial parse failed)
            try:
                base_dt = ord_dt
                if base_dt is None:
                    base_dt = pd.to_datetime(row.get('ordered_date'), dayfirst=True, errors='coerce')
                if pd.notna(base_dt):
                    # TikTok: +1 day, Lazada: +1 day, Zalora: +2 days
                    if channel == 'Zalora':
                        days = 2
                    else:
                        days = 1
                    mp_sla = (base_dt + timedelta(days=days)).strftime('%d/%m/%Y %H:%M')
                else:
                    mp_sla = ''
            except Exception:
                mp_sla = ''
        else:
            mp_sla = ''

        report_rows.append({
            'Brand':           row['Brand'],
            'Channel':         channel,
            'Nickname':        row['Nickname'],
            'Order ID':        oid,
            'FYW SKU':         row['custom_sku'],
            'Product Name':    row['item_title'],
            'Variant':         row['variant_details'],
            'Qty':             int(row['quantity']),
            'Sold Price (MYR)': round(float(row['item_sold_amount']), 2),
            'Order Date':      ord_date,
            'MP SLA':          mp_sla,
            'FYW SLA':         sla_date,
            'FYW Status':      row['order_status'],
            'Tracking No.':    tracking,
            'Courier':         str(row['courier_name']).strip().replace('\\r', '') if pd.notna(row['courier_name']) else '',
        })

    df_report = pd.DataFrame(report_rows)
    if df_report.empty:
        raise ValueError("No pending orders found in the uploaded CSV.")

    df_report['_s'] = df_report['Brand'].map(BRAND_ORDER)
    df_report = df_report.sort_values(['_s', 'Channel', 'Order Date']).drop(columns=['_s']).reset_index(drop=True)

    brand_summary = df_report.groupby('Brand').agg(
        Orders=('Order ID', 'nunique'),
        Items=('Qty', 'sum'),
        Total_Value=('Sold Price (MYR)', 'sum')
    ).reset_index()
    channel_map = df_report.groupby('Brand')['Channel'].apply(lambda x: ', '.join(sorted(set(x)))).to_dict()

    # ── Pivot ─────────────────────────────────────────────────────────────────
    pivot = df_report.pivot_table(
        index='Nickname', columns='FYW SLA',
        values='Order ID', aggfunc='count', fill_value=0
    )
    pivot['Grand Total'] = pivot.sum(axis=1)
    total_row = pivot.sum(axis=0); total_row.name = 'Grand Total'
    pivot = pd.concat([pivot, total_row.to_frame().T]).astype(int)

    nickname_brand = df_report.drop_duplicates('Nickname').set_index('Nickname')['Brand']
    def nick_sort(n):
        if n == 'Grand Total': return (99, '')
        return (BRAND_ORDER.get(nickname_brand.get(n, 'OTHER'), 3), n)
    pivot = pivot.loc[sorted(pivot.index, key=nick_sort)]

    # Sort SLA date columns chronologically (columns are strings in 'DD/MM/YYYY' format)
    def sort_date_col(col_name):
        try:
            return pd.to_datetime(col_name, format='%d/%m/%Y')
        except Exception:
            return pd.Timestamp.max
    sla_cols = sorted([c for c in pivot.columns if c != 'Grand Total'], key=sort_date_col)
    all_cols = sla_cols + ['Grand Total']

    # ─────────────────────────────────────────────────────────────────────────
    # BUILD EXCEL
    # ─────────────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── SHEET 1: PIVOT ────────────────────────────────────────────────────────
    ws_p = wb.active; ws_p.title = 'PIVOT'
    ws_p.sheet_view.showGridLines = False
    total_cols = len(all_cols) + 1

    ws_p.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws_p['A1'].value     = f'PENDING ORDERS BY NICKNAME & FYW SLA — {today.strftime("%d %B %Y")}'
    ws_p['A1'].font      = Font(name='Arial', bold=True, size=13, color='1F4E79')
    ws_p['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_p.row_dimensions[1].height = 30

    ws_p.merge_cells(f'A2:{get_column_letter(total_cols)}2')
    ws_p['A2'].value     = '🔴 Past due  |  🟠 1 day before today  |  🟢 Today  |  Values = Count of Orders'
    ws_p['A2'].font      = Font(name='Arial', italic=True, size=9, color='595959')
    ws_p['A2'].alignment = Alignment(horizontal='center')
    ws_p.row_dimensions[2].height = 16
    ws_p.row_dimensions[3].height = 8

    hdr_style(ws_p.cell(row=4, column=1, value='Nickname'), size=10)
    for col_i, col_name in enumerate(all_cols, 2):
        hdr_style(ws_p.cell(row=4, column=col_i, value=col_name),
                  bg=col_header_bg(col_name, today, yesterday), size=10)
    ws_p.row_dimensions[4].height = 24

    nicknames = [r for r in pivot.index if r != 'Grand Total']
    for row_i, nick in enumerate(nicknames, 5):
        bg = NICK_BG.get(nick, '555555')
        c = ws_p.cell(row=row_i, column=1, value=nick)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = border
        for col_i, col_name in enumerate(all_cols, 2):
            val = int(pivot.loc[nick, col_name])
            is_bold = classify_col(col_name, today, yesterday) in ('past','yesterday','today','grand_total')
            c = ws_p.cell(row=row_i, column=col_i, value=val if val > 0 else '-')
            c.font  = Font(name='Arial', bold=is_bold,
                           color=col_data_font_color(col_name, today, yesterday), size=10)
            c.fill  = PatternFill('solid', start_color=col_data_bg(col_name, today, yesterday))
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        ws_p.row_dimensions[row_i].height = 22

    grand_row = len(nicknames) + 5
    for col_i, col_name in enumerate(['Nickname'] + all_cols, 1):
        val = 'Grand Total' if col_i == 1 else int(pivot.loc['Grand Total', col_name])
        c = ws_p.cell(row=grand_row, column=col_i, value=val)
        c.font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill  = PatternFill('solid', start_color=col_grand_bg(col_name, today, yesterday) if col_i > 1 else '2E75B6')
        c.alignment = Alignment(horizontal='left' if col_i == 1 else 'center', vertical='center')
        c.border = border
    ws_p.row_dimensions[grand_row].height = 24
    ws_p.column_dimensions['A'].width = 28
    for i in range(2, len(all_cols) + 2):
        ws_p.column_dimensions[get_column_letter(i)].width = 18

    # ── SHEET 2: SUMMARY ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('SUMMARY')
    ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells('A1:G1')
    ws_sum['A1'].value     = f'FYW PENDING ORDERS REPORT — {today.strftime("%d %B %Y")}'
    ws_sum['A1'].font      = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 30

    ws_sum.merge_cells('A2:G2')
    ws_sum['A2'].value     = 'Summary as of report date | Source: FYW Dashboard + Shopee/TikTok/Zalora/Lazada Seller Centre'
    ws_sum['A2'].font      = Font(name='Arial', italic=True, size=9, color='595959')
    ws_sum['A2'].alignment = Alignment(horizontal='center')
    ws_sum.row_dimensions[2].height = 16
    ws_sum.row_dimensions[3].height = 8

    for col, h in enumerate(['Brand','Unique Orders','Total Items','Total Value (MYR)','Channels'], 1):
        hdr_style(ws_sum.cell(row=4, column=col, value=h))
    ws_sum.row_dimensions[4].height = 22

    total_orders = total_items = 0; total_val = 0.0; last_i = 5
    for i, row in enumerate(brand_summary.itertuples(), 5):
        bg = BRAND_COLORS.get(row.Brand, 'FFFFFF')
        for col, val in enumerate([row.Brand, row.Orders, row.Items, row.Total_Value, channel_map.get(row.Brand,'')], 1):
            c = ws_sum.cell(row=i, column=col, value=val)
            c.font      = Font(name='Arial', bold=(col==1), size=9)
            c.fill      = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='left' if col==1 else 'center', vertical='center')
            c.border    = border
            if col == 4: c.number_format = '#,##0.00'
        total_orders += row.Orders; total_items += row.Items; total_val += row.Total_Value
        ws_sum.row_dimensions[i].height = 18; last_i = i

    for col, val in enumerate(['TOTAL', total_orders, total_items, total_val, ''], 1):
        c = ws_sum.cell(row=last_i+1, column=col, value=val)
        c.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill      = PatternFill('solid', start_color='1F4E79')
        c.alignment = Alignment(horizontal='left' if col==1 else 'center', vertical='center')
        c.border    = border
        if col == 4: c.number_format = '#,##0.00'
    ws_sum.row_dimensions[last_i+1].height = 20
    for col, w in zip(range(1,6), [18,18,14,22,30]):
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 3: ALL PENDING ORDERS ───────────────────────────────────────────
    ws = wb.create_sheet('ALL PENDING ORDERS')
    ws.sheet_view.showGridLines = False

    cols = ['Brand','Channel','Order ID','FYW SKU','Product Name','Variant',
            'Qty','Sold Price (MYR)','Order Date','MP SLA','FYW SLA',
            'FYW Status','Tracking No.','Courier']
    CENTER_COLS = {1,2,7,8,9,10,11,12}

    ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
    ws['A1'].value     = f'FYW PENDING ORDERS — {today.strftime("%d %B %Y")} | {len(df_report)} items | MYR {df_report["Sold Price (MYR)"].sum():,.2f}'
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(cols, 1):
        hdr_style(ws.cell(row=2, column=col, value=h))
    ws.cell(row=2, column=cols.index('MP SLA')+1).value = 'MP SLA\n(Shopee=Seller Centre\nTikTok/Lazada=+1d, Zalora=+3d)'
    ws.row_dimensions[2].height = 36

    for r_idx, row in df_report.iterrows():
        er     = r_idx + 3
        row_bg = STATUS_COLORS.get(row['FYW Status'], 'FFFFFF')
        for col_idx, col_name in enumerate(cols, 1):
            c = ws.cell(row=er, column=col_idx, value=row[col_name])
            c.font      = Font(name='Arial', size=9, bold=(col_idx in [1,4]))
            c.fill      = PatternFill('solid', start_color=row_bg)
            c.alignment = Alignment(horizontal='center' if col_idx in CENTER_COLS else 'left',
                                    vertical='center', wrap_text=True)
            c.border    = border
            if col_idx == 8: c.number_format = '#,##0.00'
        ws.row_dimensions[er].height = 30

    for i, w in enumerate([12,10,22,22,44,28,6,18,13,20,20,20,25,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    # ── SHEETS 4+: PER BRAND ──────────────────────────────────────────────────
    cols_b   = ['Channel','Order ID','FYW SKU','Product Name','Variant',
                'Qty','Sold Price (MYR)','Order Date','MP SLA','FYW SLA',
                'FYW Status','Tracking No.','Courier']
    CENTER_B = {1,6,7,8,9,10,11}

    for brand in ['MELISSA','IPANEMA','CSPACE']:
        df_b = df_report[df_report['Brand'] == brand].copy()
        if df_b.empty: continue
        ws_b   = wb.create_sheet(brand)
        ws_b.sheet_view.showGridLines = False
        hdr_bg = SUBHEADER_BG.get(brand, '1F4E79')

        ws_b.merge_cells(f'A1:{get_column_letter(len(cols_b))}1')
        ws_b['A1'].value     = f'{brand} — PENDING ORDERS | {today.strftime("%d %B %Y")} | {len(df_b)} items | MYR {df_b["Sold Price (MYR)"].sum():,.2f}'
        ws_b['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws_b['A1'].fill      = PatternFill('solid', start_color=hdr_bg)
        ws_b['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_b.row_dimensions[1].height = 28

        for col, h in enumerate(cols_b, 1):
            hdr_style(ws_b.cell(row=2, column=col, value=h), bg=hdr_bg)
        ws_b.cell(row=2, column=cols_b.index('MP SLA')+1).value = 'MP SLA\n(Shopee=Seller Centre\nTikTok/Lazada=+1d, Zalora=+3d)'
        ws_b.row_dimensions[2].height = 36

        for r_idx, (_, row) in enumerate(df_b.iterrows()):
            er     = r_idx + 3
            row_bg = STATUS_COLORS.get(row['FYW Status'], BRAND_COLORS.get(brand,'FFFFFF'))
            for col_idx, col_name in enumerate(cols_b, 1):
                c = ws_b.cell(row=er, column=col_idx, value=row[col_name])
                c.font      = Font(name='Arial', size=9, bold=(col_idx==3))
                c.fill      = PatternFill('solid', start_color=row_bg)
                c.alignment = Alignment(horizontal='center' if col_idx in CENTER_B else 'left',
                                        vertical='center', wrap_text=True)
                c.border    = border
                if col_idx == 7: c.number_format = '#,##0.00'
            ws_b.row_dimensions[er].height = 30

        tot_row = len(df_b) + 3
        for col in range(1, len(cols_b)+1):
            c = ws_b.cell(row=tot_row, column=col, value='')
            c.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill      = PatternFill('solid', start_color=hdr_bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws_b.cell(row=tot_row, column=1).value     = 'TOTAL'
        ws_b.cell(row=tot_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws_b.cell(row=tot_row, column=6).value     = f'=SUM(F3:F{tot_row-1})'
        ws_b.cell(row=tot_row, column=7).value     = f'=SUM(G3:G{tot_row-1})'
        ws_b.cell(row=tot_row, column=7).number_format = '#,##0.00'
        ws_b.row_dimensions[tot_row].height = 20

        for i, w in enumerate([10,22,22,44,28,6,18,13,20,20,20,25,18], 1):
            ws_b.column_dimensions[get_column_letter(i)].width = w
        ws_b.freeze_panes = 'A3'

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    summary = {
        'total_items':   len(df_report),
        'total_value':   round(df_report['Sold Price (MYR)'].sum(), 2),
        'nicknames':     df_report['Nickname'].unique().tolist(),
        'sla_dates':     sla_cols,
        'pivot_df':      pivot,
        'warnings':      load_warnings,
        'report_date':   today.strftime('%d %B %Y'),
    }
    return buffer.getvalue(), summary
