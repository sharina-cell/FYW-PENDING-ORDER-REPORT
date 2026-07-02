"""
FYW Pending Order Report — Core Engine
Generates a formatted Excel workbook from FYW dashboard + Shopee/TikTok/Zalora/Lazada files.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from io import BytesIO


# ── Constants ────────────────────────────────────────────────────────────────

PENDING_STATUSES = ['ACCEPTED/PICKED', 'READY TO SHIP', 'NEW', 'CANCEL REQUESTED']

BRAND_COLORS  = {'MELISSA': 'D6E4F7', 'IPANEMA': 'D6F0E0', 'CSPACE': 'FFF2CC'}
HEADER_BG     = '1F4E79'
SUBHEADER_BG  = {'MELISSA': '2E75B6', 'IPANEMA': '375623', 'CSPACE': 'BF8F00'}
STATUS_COLORS = {
    'ACCEPTED/PICKED':  'E2EFDA',
    'READY TO SHIP':    'FFF2CC',
    'NEW':              'DDEEFF',
    'CANCEL REQUESTED': 'FCE4D6',
}
NICK_BG = {
    'shopee-Melissa MY':        '2E75B6',
    'shopee-Ipanema MY':        '375623',
    'shopee-CSpace MY':         'BF8F00',
    'tiktok-tiktok-CspaceMY':   '444444',
    'tiktok-MelissaMY':         '444444',
    'tiktok-tiktok-IpanemaMY':  '444444',
    'zalora-Melissa-MY':        'C00000',
    'zalora-Ipanema-MY':        'C00000',
    'lazada-Melissa MY':        '0F146D',
    'lazada-Ipanema MY':        '0F146D',
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def get_brand(nickname: str) -> str:
    if pd.isna(nickname):
        return 'OTHER'
    n = str(nickname).lower()
    if 'melissa' in n: return 'MELISSA'
    if 'ipanema' in n: return 'IPANEMA'
    if 'cspace'  in n: return 'CSPACE'
    return 'OTHER'


def get_channel(nickname: str) -> str:
    if pd.isna(nickname):
        return ''
    n = str(nickname).lower()
    if 'shopee' in n: return 'Shopee'
    if 'tiktok' in n: return 'TikTok'
    if 'lazada' in n: return 'Lazada'
    if 'zalora' in n: return 'Zalora'
    return str(nickname)


def classify_col(col_name: str, today: datetime.date, yesterday: datetime.date) -> str:
    """Returns: past | yesterday | today | future | grand_total"""
    if col_name == 'Grand Total':
        return 'grand_total'
    try:
        col_date = datetime.strptime(col_name, '%d/%m/%Y').date()
        if col_date < yesterday:  return 'past'
        if col_date == yesterday: return 'yesterday'
        if col_date == today:     return 'today'
        return 'future'
    except Exception:
        return 'future'


def col_header_bg(col_name, today, yesterday):
    c = classify_col(col_name, today, yesterday)
    return {'past':'C00000','yesterday':'C55A00','today':'375623','grand_total':'2E75B6'}.get(c, '1F4E79')


def col_data_bg(col_name, today, yesterday):
    c = classify_col(col_name, today, yesterday)
    return {'past':'FFCCCC','yesterday':'FCE4D6','today':'C6EFCE','grand_total':'DDEEFF'}.get(c, 'F2F2F2')


def col_data_font_color(col_name, today, yesterday):
    c = classify_col(col_name, today, yesterday)
    return {'past':'9C0006','yesterday':'833C00','today':'006100'}.get(c, '000000')


def col_grand_bg(col_name, today, yesterday):
    c = classify_col(col_name, today, yesterday)
    return {'past':'C00000','yesterday':'C55A00','today':'70AD47','grand_total':'2E75B6'}.get(c, '1F4E79')


def make_border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr_style(cell, bg=HEADER_BG, fg='FFFFFF', size=9):
    border = make_border()
    cell.font      = Font(name='Arial', bold=True, color=fg, size=size)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = border


# ── Data Loading ─────────────────────────────────────────────────────────────

def load_fyw_csv(file) -> pd.DataFrame:
    """Load and filter FYW CSV to pending orders only."""
    df = pd.read_csv(file)
    df['order_id_clean']  = df['order_id'].str.strip()
    df['tracking_clean']  = df['tracking_number'].str.strip()
    pending = df[df['order_status'].isin(PENDING_STATUSES)].copy()
    pending['order_id_clean'] = pending['order_id'].str.strip()
    pending['tracking_clean'] = pending['tracking_number'].str.strip()
    pending['Nickname']   = pending['nickname'].str.strip()
    pending['Brand']      = pending['nickname'].apply(get_brand)
    pending['Channel']    = pending['nickname'].apply(get_channel)
    return pending


def load_shopee_files(shopee_file_map: dict) -> tuple[dict, dict]:
    """
    Load Shopee Excel files and build lookup maps.
    shopee_file_map: {'MELISSA': file_object, 'IPANEMA': file_object, 'CSPACE': file_object}
    Returns: (mp_sla_map, tracking_map)
    """
    mp_sla_map      = {}
    shopee_tracking = {}
    for brand, file_obj in shopee_file_map.items():
        if file_obj is None:
            continue
        try:
            df = pd.read_excel(file_obj)
            for _, row in df.iterrows():
                oid = str(row['Order ID']).strip()
                est = row.get('Estimated Ship Out Date', None)
                if pd.notna(est) and str(est).strip() not in ['', 'nan']:
                    mp_sla_map[oid] = str(est).strip()
                trk = str(row['Tracking Number*']).strip() if pd.notna(row.get('Tracking Number*')) else ''
                if trk and trk != 'nan':
                    shopee_tracking[oid] = trk
        except Exception as e:
            print(f'⚠️  Could not load {brand} Shopee file: {e}')
    return mp_sla_map, shopee_tracking


# ── Report Building ──────────────────────────────────────────────────────────

def build_report_rows(pending: pd.DataFrame, mp_sla_map: dict, shopee_tracking: dict) -> pd.DataFrame:
    """Build the flat report dataframe from pending orders."""
    rows = []
    for _, row in pending.iterrows():
        oid     = row['order_id_clean']
        channel = row['Channel']

        # Tracking
        tracking = row['tracking_clean']
        if not tracking or tracking == 'nan':
            tracking = shopee_tracking.get(oid, '')

        # Order date
        try:
            ord_dt   = pd.to_datetime(row['ordered_date'], dayfirst=True)
            ord_date = ord_dt.strftime('%d/%m/%Y')
        except Exception:
            ord_dt   = None
            ord_date = str(row['ordered_date'])

        # FYW SLA
        try:
            sla_date = pd.to_datetime(row['order_sla'], dayfirst=True).strftime('%d/%m/%Y')
        except Exception:
            sla_date = str(row['order_sla']) if pd.notna(row['order_sla']) else ''

        # MP SLA
        if channel == 'Shopee':
            raw = mp_sla_map.get(oid, '')
            try:
                mp_sla = pd.to_datetime(raw).strftime('%d/%m/%Y %H:%M') if raw else ''
            except Exception:
                mp_sla = raw
        elif channel == 'TikTok' and ord_dt:
            mp_sla = (ord_dt + timedelta(days=1)).strftime('%d/%m/%Y %H:%M')
        elif channel == 'Zalora' and ord_dt:
            mp_sla = (ord_dt + timedelta(days=3)).strftime('%d/%m/%Y %H:%M')
        elif channel == 'Lazada' and ord_dt:
            mp_sla = (ord_dt + timedelta(days=1)).strftime('%d/%m/%Y %H:%M')
        else:
            mp_sla = ''

        rows.append({
            'Brand':            row['Brand'],
            'Channel':          channel,
            'Nickname':         row['Nickname'],
            'Order ID':         oid,
            'FYW SKU':          row['custom_sku'],
            'Product Name':     row['item_title'],
            'Variant':          row['variant_details'],
            'Qty':              int(row['quantity']),
            'Sold Price (MYR)': round(float(row['item_sold_amount']), 2),
            'Order Date':       ord_date,
            'MP SLA':           mp_sla,
            'FYW SLA':          sla_date,
            'FYW Status':       row['order_status'],
            'Tracking No.':     tracking,
            'Courier':          str(row['courier_name']).strip().replace('\\r', '')
                                if pd.notna(row['courier_name']) else '',
        })

    BMAP = {'MELISSA': 0, 'IPANEMA': 1, 'CSPACE': 2, 'OTHER': 3}
    df = pd.DataFrame(rows)
    df['Invoice Number']  = ''
    df['Payment Status']  = ''
    df['_s'] = df['Brand'].map(BMAP)
    df = df.sort_values(['_s', 'Channel', 'Order Date']).drop(columns=['_s']).reset_index(drop=True)
    return df


def build_pivot(df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """Build pivot table: rows=Nickname, columns=FYW SLA, values=order count."""
    BMAP = {'MELISSA': 0, 'IPANEMA': 1, 'CSPACE': 2, 'OTHER': 3}
    pivot = df.pivot_table(index='Nickname', columns='FYW SLA', values='Order ID',
                           aggfunc='count', fill_value=0)
    pivot['Grand Total'] = pivot.sum(axis=1)
    total_row = pivot.sum(axis=0)
    total_row.name = 'Grand Total'
    pivot = pd.concat([pivot, total_row.to_frame().T]).astype(int)

    nickname_brand = df.drop_duplicates('Nickname').set_index('Nickname')['Brand']

    def nick_sort(n):
        if n == 'Grand Total': return (99, '')
        return (BMAP.get(nickname_brand.get(n, 'OTHER'), 3), n)

    pivot = pivot.loc[sorted(pivot.index, key=nick_sort)]
    sla_cols = [c for c in pivot.columns if c != 'Grand Total']
    all_cols  = sla_cols + ['Grand Total']
    return pivot, all_cols


# ── Manual Reconciliation Merge ─────────────────────────────────────────────

def merge_manual_orders(df: pd.DataFrame, manual_confirmed: dict) -> pd.DataFrame:
    """
    Fold manually-confirmed "pushed to TC" orders (captured in the TC
    Reconciliation Check tab) into the main report dataframe, so they appear
    in FILTERED DATA / PIVOT just like any other pending order.

    manual_confirmed: { marketplace_label: [ {eOrder Number, Invoice Number,
        Payment Status, Order Item Status, Ordered Date, Nickname, MP SLA}, ... ] }
    """
    new_rows = []
    for label, entries in manual_confirmed.items():
        for e in entries:
            nickname = e.get('Nickname') or ''
            new_rows.append({
                'Brand':            get_brand(nickname) if nickname else 'OTHER',
                'Channel':          get_channel(nickname) if nickname else label,
                'Nickname':         nickname,
                'Order ID':         e.get('eOrder Number', ''),
                'FYW SKU':          '',
                'Product Name':     '(Manually confirmed — not caught by auto-match)',
                'Variant':          '',
                'Qty':              0,
                'Sold Price (MYR)': 0.0,
                'Order Date':       e.get('Ordered Date', ''),
                'MP SLA':           e.get('MP SLA', ''),
                'FYW SLA':          '',
                'FYW Status':       e.get('Order Item Status', ''),
                'Tracking No.':     '',
                'Courier':          '',
                'Invoice Number':   e.get('Invoice Number', ''),
                'Payment Status':   e.get('Payment Status', ''),
            })

    if not new_rows:
        return df.copy()

    manual_df = pd.DataFrame(new_rows)
    combined = pd.concat([df, manual_df], ignore_index=True, sort=False)

    BMAP = {'MELISSA': 0, 'IPANEMA': 1, 'CSPACE': 2, 'OTHER': 3}
    combined['_s'] = combined['Brand'].map(BMAP).fillna(3)
    combined = combined.sort_values(['_s', 'Channel', 'Order Date']).drop(columns=['_s']).reset_index(drop=True)
    return combined


# ── NOT PUSHED Sheet ─────────────────────────────────────────────────────────

def add_not_pushed_sheet(wb, results: dict, border) -> None:
    """
    Add a consolidated 'NOT PUSHED' sheet listing every order still missing
    from TC across all checked marketplaces (i.e. not yet applied via manual
    confirmation and not caught by an auto-match).
    """
    rows = []
    for label, res in results.items():
        if 'error' in res:
            continue
        for oid in res.get('missing_ids', []):
            rows.append((label, oid))

    ws = wb.create_sheet('NOT PUSHED')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:C1')
    ws['A1'].value     = f'ORDERS NOT PUSHED TO TC ({len(rows)} total)'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', start_color='C00000')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(['#', 'Marketplace', 'Order ID / Order Number'], 1):
        hdr_style(ws.cell(row=2, column=col, value=h), bg='C00000')
    ws.row_dimensions[2].height = 22

    for i, (label, oid) in enumerate(rows, 1):
        er = i + 2
        vals = [i, label, oid]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=er, column=col, value=val)
            c.font      = Font(name='Arial', size=9)
            c.fill      = PatternFill('solid', start_color='FFCCCC')
            c.alignment = Alignment(horizontal='center' if col in (1, 2) else 'left', vertical='center')
            c.border    = border
        ws.row_dimensions[er].height = 18

    if not rows:
        ws.merge_cells('A3:C3')
        ws['A3'].value     = '🎉 All marketplace orders have been pushed to TC.'
        ws['A3'].font      = Font(name='Arial', italic=True, size=10, color='006100')
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 22

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 30
    ws.freeze_panes = 'A3'


# ── Excel Writer ─────────────────────────────────────────────────────────────

def generate_excel(df: pd.DataFrame, report_date: datetime = None, not_pushed_results: dict = None) -> BytesIO:
    """
    Generate the full Excel workbook and return as a BytesIO buffer.
    report_date defaults to today if not provided.

    not_pushed_results (optional): the TC-reconciliation `results` dict
    (see tc_reconciliation.run_full_reconciliation / apply_manual_overrides).
    If provided, an extra 'NOT PUSHED' sheet is added listing every order
    still missing from TC.
    """
    if report_date is None:
        report_date = datetime.today()

    today     = report_date.date()
    yesterday = today - timedelta(days=1)
    BMAP      = {'MELISSA': 0, 'IPANEMA': 1, 'CSPACE': 2, 'OTHER': 3}
    border    = make_border()

    pivot, all_cols = build_pivot(df)
    sla_cols = [c for c in all_cols if c != 'Grand Total']

    brand_summary = df.groupby('Brand').agg(
        Orders=('Order ID', 'nunique'),
        Items=('Qty', 'sum'),
        Total_Value=('Sold Price (MYR)', 'sum')
    ).reset_index()
    channel_map = df.groupby('Brand')['Channel'].apply(
        lambda x: ', '.join(sorted(set(x)))
    ).to_dict()

    wb = openpyxl.Workbook()

    # ── SHEET 1: PIVOT ────────────────────────────────────────────────────────
    ws_p = wb.active
    ws_p.title = 'PIVOT'
    ws_p.sheet_view.showGridLines = False

    total_cols = len(all_cols) + 1
    ws_p.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws_p['A1'].value     = f'PENDING ORDERS BY NICKNAME & FYW SLA — {report_date.strftime("%d %B %Y")}'
    ws_p['A1'].font      = Font(name='Arial', bold=True, size=13, color='1F4E79')
    ws_p['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_p.row_dimensions[1].height = 30

    ws_p.merge_cells(f'A2:{get_column_letter(total_cols)}2')
    ws_p['A2'].value     = '🔴 Past due  |  🟠 1 day before today  |  🟢 Today  |  Values = Count of Orders'
    ws_p['A2'].font      = Font(name='Arial', italic=True, size=9, color='595959')
    ws_p['A2'].alignment = Alignment(horizontal='center')
    ws_p.row_dimensions[2].height = 16
    ws_p.row_dimensions[3].height = 8

    hdr_style(ws_p.cell(row=4, column=1, value='Nickname'), size=10)
    for col_i, col_name in enumerate(all_cols, 2):
        hdr_style(ws_p.cell(row=4, column=col_i, value=col_name),
                  bg=col_header_bg(col_name, today, yesterday), size=10)
    ws_p.row_dimensions[4].height = 24

    nicknames = [r for r in pivot.index if r != 'Grand Total']
    for row_i, nick in enumerate(nicknames, 5):
        bg = NICK_BG.get(nick, '555555')
        c  = ws_p.cell(row=row_i, column=1, value=nick)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill      = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = border

        for col_i, col_name in enumerate(all_cols, 2):
            val     = int(pivot.loc[nick, col_name])
            is_bold = classify_col(col_name, today, yesterday) in ('past', 'yesterday', 'today', 'grand_total')
            c = ws_p.cell(row=row_i, column=col_i, value=val if val > 0 else '-')
            c.font      = Font(name='Arial', bold=is_bold,
                               color=col_data_font_color(col_name, today, yesterday), size=10)
            c.fill      = PatternFill('solid', start_color=col_data_bg(col_name, today, yesterday))
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws_p.row_dimensions[row_i].height = 22

    grand_row = len(nicknames) + 5
    for col_i, col_name in enumerate(['Nickname'] + all_cols, 1):
        val = 'Grand Total' if col_i == 1 else int(pivot.loc['Grand Total', col_name])
        c   = ws_p.cell(row=grand_row, column=col_i, value=val)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill      = PatternFill('solid', start_color=col_grand_bg(col_name, today, yesterday)
                                  if col_i > 1 else '2E75B6')
        c.alignment = Alignment(horizontal='left' if col_i == 1 else 'center', vertical='center')
        c.border    = border
    ws_p.row_dimensions[grand_row].height = 24

    ws_p.column_dimensions['A'].width = 28
    for i in range(2, len(all_cols) + 2):
        ws_p.column_dimensions[get_column_letter(i)].width = 18

    # ── SHEET 2: SUMMARY ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('SUMMARY')
    ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells('A1:G1')
    ws_sum['A1'].value     = f'FYW PENDING ORDERS REPORT — {report_date.strftime("%d %B %Y")}'
    ws_sum['A1'].font      = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 30

    ws_sum.merge_cells('A2:G2')
    ws_sum['A2'].value     = 'Summary as of report date | Source: FYW Dashboard + Shopee/TikTok/Zalora/Lazada'
    ws_sum['A2'].font      = Font(name='Arial', italic=True, size=9, color='595959')
    ws_sum['A2'].alignment = Alignment(horizontal='center')
    ws_sum.row_dimensions[2].height = 16
    ws_sum.row_dimensions[3].height = 8

    for col, h in enumerate(['Brand', 'Unique Orders', 'Total Items', 'Total Value (MYR)', 'Channels'], 1):
        hdr_style(ws_sum.cell(row=4, column=col, value=h))
    ws_sum.row_dimensions[4].height = 22

    total_orders = total_items = 0
    total_val = 0.0
    last_i = 5
    for i, row in enumerate(brand_summary.itertuples(), 5):
        bg = BRAND_COLORS.get(row.Brand, 'FFFFFF')
        for col, val in enumerate(
            [row.Brand, row.Orders, row.Items, row.Total_Value, channel_map.get(row.Brand, '')], 1
        ):
            c = ws_sum.cell(row=i, column=col, value=val)
            c.font      = Font(name='Arial', bold=(col == 1), size=9)
            c.fill      = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')
            c.border    = border
            if col == 4:
                c.number_format = '#,##0.00'
        total_orders += row.Orders
        total_items  += row.Items
        total_val    += row.Total_Value
        ws_sum.row_dimensions[i].height = 18
        last_i = i

    for col, val in enumerate(['TOTAL', total_orders, total_items, total_val, ''], 1):
        c = ws_sum.cell(row=last_i + 1, column=col, value=val)
        c.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill      = PatternFill('solid', start_color='1F4E79')
        c.alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')
        c.border    = border
        if col == 4:
            c.number_format = '#,##0.00'
    ws_sum.row_dimensions[last_i + 1].height = 20

    for col, w in zip(range(1, 6), [18, 18, 14, 22, 30]):
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 3: ALL PENDING ORDERS ───────────────────────────────────────────
    ws = wb.create_sheet('ALL PENDING ORDERS')
    ws.sheet_view.showGridLines = False

    DETAIL_COLS  = ['Brand', 'Channel', 'Order ID', 'FYW SKU', 'Product Name', 'Variant',
                    'Qty', 'Sold Price (MYR)', 'Order Date', 'MP SLA', 'FYW SLA',
                    'FYW Status', 'Tracking No.', 'Courier', 'Invoice Number', 'Payment Status']
    CENTER_COLS  = {1, 2, 7, 8, 9, 10, 11, 12}
    COL_WIDTHS   = [12, 10, 22, 22, 44, 28, 6, 18, 13, 20, 20, 20, 25, 18, 18, 16]

    ws.merge_cells(f'A1:{get_column_letter(len(DETAIL_COLS))}1')
    ws['A1'].value     = (f'FYW PENDING ORDERS — {report_date.strftime("%d %B %Y")} | '
                          f'{len(df)} items | MYR {df["Sold Price (MYR)"].sum():,.2f}')
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(DETAIL_COLS, 1):
        hdr_style(ws.cell(row=2, column=col, value=h))
    ws.cell(row=2, column=DETAIL_COLS.index('MP SLA') + 1).value = \
        'MP SLA\n(Shopee=Seller Centre\nTikTok/Lazada=+1d, Zalora=+3d)'
    ws.row_dimensions[2].height = 36

    for r_idx, row in df.iterrows():
        er     = r_idx + 3
        row_bg = STATUS_COLORS.get(row['FYW Status'], 'FFFFFF')
        for col_idx, col_name in enumerate(DETAIL_COLS, 1):
            c = ws.cell(row=er, column=col_idx, value=row[col_name])
            c.font      = Font(name='Arial', size=9, bold=(col_idx in [1, 4]))
            c.fill      = PatternFill('solid', start_color=row_bg)
            c.alignment = Alignment(
                horizontal='center' if col_idx in CENTER_COLS else 'left',
                vertical='center', wrap_text=True
            )
            c.border = border
            if col_idx == 8:
                c.number_format = '#,##0.00'
        ws.row_dimensions[er].height = 30

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    # ── SHEETS 4+: PER BRAND ──────────────────────────────────────────────────
    BRAND_COLS   = ['Channel', 'Order ID', 'FYW SKU', 'Product Name', 'Variant',
                    'Qty', 'Sold Price (MYR)', 'Order Date', 'MP SLA', 'FYW SLA',
                    'FYW Status', 'Tracking No.', 'Courier', 'Invoice Number', 'Payment Status']
    CENTER_B     = {1, 6, 7, 8, 9, 10, 11}
    BRAND_WIDTHS = [10, 22, 22, 44, 28, 6, 18, 13, 20, 20, 20, 25, 18, 18, 16]

    for brand in ['MELISSA', 'IPANEMA', 'CSPACE']:
        df_b = df[df['Brand'] == brand].copy()
        if df_b.empty:
            continue

        ws_b   = wb.create_sheet(brand)
        ws_b.sheet_view.showGridLines = False
        hdr_bg = SUBHEADER_BG.get(brand, '1F4E79')

        ws_b.merge_cells(f'A1:{get_column_letter(len(BRAND_COLS))}1')
        ws_b['A1'].value     = (f'{brand} — PENDING ORDERS | {report_date.strftime("%d %B %Y")} | '
                                f'{len(df_b)} items | MYR {df_b["Sold Price (MYR)"].sum():,.2f}')
        ws_b['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws_b['A1'].fill      = PatternFill('solid', start_color=hdr_bg)
        ws_b['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_b.row_dimensions[1].height = 28

        for col, h in enumerate(BRAND_COLS, 1):
            hdr_style(ws_b.cell(row=2, column=col, value=h), bg=hdr_bg)
        ws_b.cell(row=2, column=BRAND_COLS.index('MP SLA') + 1).value = \
            'MP SLA\n(Shopee=Seller Centre\nTikTok/Lazada=+1d, Zalora=+3d)'
        ws_b.row_dimensions[2].height = 36

        for r_idx, (_, row) in enumerate(df_b.iterrows()):
            er     = r_idx + 3
            row_bg = STATUS_COLORS.get(row['FYW Status'], BRAND_COLORS.get(brand, 'FFFFFF'))
            for col_idx, col_name in enumerate(BRAND_COLS, 1):
                c = ws_b.cell(row=er, column=col_idx, value=row[col_name])
                c.font      = Font(name='Arial', size=9, bold=(col_idx == 3))
                c.fill      = PatternFill('solid', start_color=row_bg)
                c.alignment = Alignment(
                    horizontal='center' if col_idx in CENTER_B else 'left',
                    vertical='center', wrap_text=True
                )
                c.border = border
                if col_idx == 7:
                    c.number_format = '#,##0.00'
            ws_b.row_dimensions[er].height = 30

        tot_row = len(df_b) + 3
        for col in range(1, len(BRAND_COLS) + 1):
            c = ws_b.cell(row=tot_row, column=col, value='')
            c.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill      = PatternFill('solid', start_color=hdr_bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws_b.cell(row=tot_row, column=1).value     = 'TOTAL'
        ws_b.cell(row=tot_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws_b.cell(row=tot_row, column=6).value     = f'=SUM(F3:F{tot_row - 1})'
        ws_b.cell(row=tot_row, column=7).value     = f'=SUM(G3:G{tot_row - 1})'
        ws_b.cell(row=tot_row, column=7).number_format = '#,##0.00'
        ws_b.row_dimensions[tot_row].height = 20

        for i, w in enumerate(BRAND_WIDTHS, 1):
            ws_b.column_dimensions[get_column_letter(i)].width = w
        ws_b.freeze_panes = 'A3'

    # ── SHEET: NOT PUSHED (optional) ─────────────────────────────────────────
    if not_pushed_results:
        add_not_pushed_sheet(wb, not_pushed_results, border)

    # ── Save to buffer ────────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
