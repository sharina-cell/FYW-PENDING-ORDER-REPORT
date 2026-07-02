"""
TC Reconciliation Module
Cross-checks marketplace order exports (Shopee, TikTok, Lazada, Zalora) against
the FYW dashboard (ALL CSV / "TC") to find orders that have NOT been pushed to TC.
"""

import copy
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# Fields captured for an order that's confirmed pushed to TC but wasn't
# caught by the automated ID match (manual override entry).
MANUAL_ENTRY_FIELDS = [
    'eOrder Number', 'Invoice Number', 'Payment Status',
    'Order Item Status', 'Ordered Date', 'Nickname', 'MP SLA',
]


# Columns that could represent the order identifier in marketplace files.
# Checked in this priority order — first match wins.
ORDER_ID_CANDIDATES = [
    'Order ID', 'order_id', 'OrderID',
    'Order Number', 'order_number', 'OrderNumber', 'Order No', 'Order No.',
]


def _find_order_id_column(df: pd.DataFrame) -> str | None:
    """Find the column in a marketplace file that represents the order ID/number."""
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for candidate in ORDER_ID_CANDIDATES:
        if candidate.lower() in cols_lower:
            return cols_lower[candidate.lower()]
    return None


def extract_mp_order_ids(file_obj, marketplace: str) -> tuple[set, str | None]:
    """
    Read a marketplace file (Shopee/TikTok/Lazada/Zalora export) and extract
    the set of order IDs/numbers it contains.

    Returns: (set_of_order_ids, column_name_used)
    """
    if file_obj is None:
        return set(), None

    try:
        if str(file_obj.name).lower().endswith('.csv'):
            df = pd.read_csv(file_obj)
        else:
            df = pd.read_excel(file_obj)
    except Exception as e:
        raise ValueError(f"Could not read {marketplace} file: {e}")

    # TikTok exports have a subheader/description row right under the real
    # header row (e.g. "Platform unique order ID."). Real data starts on the
    # row after that, so drop it before extracting IDs.
    if 'tiktok' in marketplace.lower() and not df.empty:
        df = df.iloc[1:].reset_index(drop=True)

    col = _find_order_id_column(df)
    if col is None:
        raise ValueError(
            f"Could not find an Order ID / Order Number column in the {marketplace} file. "
            f"Available columns: {list(df.columns)}"
        )

    ids = (
        df[col]
        .dropna()
        .astype(str)
        .str.strip()
        .replace('', pd.NA)
        .dropna()
        .unique()
    )
    return set(ids), col


def get_tc_order_ids(fyw_df: pd.DataFrame) -> set:
    """
    Extract the full set of order IDs and order numbers present in the
    FYW dashboard (TC) CSV — matches on either field so we don't miss orders
    that only populate one of the two columns.
    """
    ids = set()
    if 'order_id' in fyw_df.columns:
        ids |= set(fyw_df['order_id'].dropna().astype(str).str.strip().unique())
    if 'order_number' in fyw_df.columns:
        ids |= set(fyw_df['order_number'].dropna().astype(str).str.strip().unique())
    ids.discard('')
    return ids


def reconcile_marketplace(file_obj, marketplace: str, tc_order_ids: set) -> dict:
    """
    Compare a single marketplace file's order IDs against the TC (FYW dashboard) order IDs.

    Returns a dict with:
        marketplace, column_used, total_mp_orders, matched_count,
        missing_count, missing_ids (sorted list), match_rate
    """
    mp_ids, col_used = extract_mp_order_ids(file_obj, marketplace)

    if not mp_ids:
        return {
            'marketplace': marketplace,
            'column_used': col_used,
            'total_mp_orders': 0,
            'matched_count': 0,
            'missing_count': 0,
            'missing_ids': [],
            'match_rate': None,
        }

    missing = sorted(mp_ids - tc_order_ids)
    matched = len(mp_ids) - len(missing)

    return {
        'marketplace': marketplace,
        'column_used': col_used,
        'total_mp_orders': len(mp_ids),
        'matched_count': matched,
        'missing_count': len(missing),
        'missing_ids': missing,
        'match_rate': round(matched / len(mp_ids) * 100, 1) if mp_ids else None,
    }


def run_full_reconciliation(fyw_df: pd.DataFrame, mp_files: dict) -> dict:
    """
    Run reconciliation across all marketplaces.

    mp_files: {
        'Shopee - Melissa': file_obj or None,
        'Shopee - Ipanema': file_obj or None,
        'Shopee - CSpace':  file_obj or None,
        'TikTok':           file_obj or None,
        'Lazada':           file_obj or None,
        'Zalora':           file_obj or None,
    }

    Returns: {marketplace_label: reconciliation_result_dict}
    """
    tc_ids = get_tc_order_ids(fyw_df)
    results = {}
    for label, file_obj in mp_files.items():
        if file_obj is None:
            continue
        try:
            results[label] = reconcile_marketplace(file_obj, label, tc_ids)
        except ValueError as e:
            results[label] = {
                'marketplace': label,
                'column_used': None,
                'total_mp_orders': 0,
                'matched_count': 0,
                'missing_count': 0,
                'missing_ids': [],
                'match_rate': None,
                'error': str(e),
            }
    return results


# ── Manual Reconciliation (orders confirmed pushed but missed by auto-match) ──

def build_manual_entry_template(missing_ids: list) -> pd.DataFrame:
    """
    Build an editable staging table for a marketplace's still-missing orders.
    Each row defaults 'eOrder Number' to the missing ID; the rest is blank for
    the user to fill in once they've confirmed the order was actually pushed.
    A 'Confirmed' checkbox marks which rows should be applied.
    """
    n = len(missing_ids)
    data = {'Confirmed': [False] * n, 'eOrder Number': list(missing_ids)}
    for field in MANUAL_ENTRY_FIELDS[1:]:
        data[field] = [''] * n
    return pd.DataFrame(data)


def apply_manual_overrides(results: dict, manual_confirmed: dict) -> dict:
    """
    Given base reconciliation `results` and a dict of
    {marketplace_label: [ {eOrder Number, Invoice Number, ...}, ... ]}
    of orders the user has manually confirmed are pushed to TC, return an
    adjusted results dict with those orders removed from missing_ids and
    counts/match_rate recalculated. Does not mutate the input.
    """
    adjusted = copy.deepcopy(results)
    for label, rows in manual_confirmed.items():
        if label not in adjusted or not rows:
            continue
        res = adjusted[label]
        if 'missing_ids' not in res:
            continue
        confirmed_ids = {r['eOrder Number'] for r in rows if r.get('eOrder Number')}
        still_missing = [oid for oid in res['missing_ids'] if oid not in confirmed_ids]
        newly_confirmed = len(res['missing_ids']) - len(still_missing)

        res['missing_ids'] = still_missing
        res['missing_count'] = len(still_missing)
        res['matched_count'] = res.get('matched_count', 0) + newly_confirmed
        res['manual_confirmed_count'] = newly_confirmed
        total = res.get('total_mp_orders', 0)
        res['match_rate'] = round(res['matched_count'] / total * 100, 1) if total else None
    return adjusted


def manual_confirmed_to_df(manual_confirmed: dict) -> pd.DataFrame:
    """Flatten the {label: [rows]} manual confirmation store into one DataFrame."""
    rows = []
    for label, entries in manual_confirmed.items():
        for r in entries:
            row = {'Marketplace': label}
            row.update({k: r.get(k, '') for k in MANUAL_ENTRY_FIELDS})
            rows.append(row)
    if not rows:
        return pd.DataFrame(columns=['Marketplace'] + MANUAL_ENTRY_FIELDS)
    return pd.DataFrame(rows)


# ── Excel Export ─────────────────────────────────────────────────────────────

def generate_reconciliation_excel(results: dict, report_date: datetime = None,
                                   manual_df: pd.DataFrame = None) -> BytesIO:
    """
    Generate a formatted Excel workbook summarizing the TC reconciliation.

    `results` should already have manual overrides applied (via
    apply_manual_overrides) if any orders were manually confirmed, so the
    summary sheet reflects the true remaining "missing from TC" count.

    `manual_df` (optional) is the flattened manual-confirmation table
    (see manual_confirmed_to_df) — if non-empty, it's added as its own sheet.
    """
    if report_date is None:
        report_date = datetime.today()

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    HEADER_BG = '1F4E79'
    GOOD_BG   = 'C6EFCE'
    BAD_BG    = 'FFCCCC'

    def hdr_style(cell, bg=HEADER_BG, fg='FFFFFF', size=10):
        cell.font      = Font(name='Arial', bold=True, color=fg, size=size)
        cell.fill      = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border

    wb = openpyxl.Workbook()

    # ── SHEET 1: SUMMARY ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'TC RECONCILIATION SUMMARY'
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:G1')
    ws['A1'].value     = f'MARKETPLACE → TC RECONCILIATION — {report_date.strftime("%d %B %Y")}'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:G2')
    ws['A2'].value     = 'Checks whether marketplace orders have been pushed to the FYW dashboard (TC)'
    ws['A2'].font      = Font(name='Arial', italic=True, size=9, color='595959')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8

    headers = ['Marketplace', 'Order ID Column Used', 'Total MP Orders', 'Matched in TC',
               'Manually Confirmed', 'Missing from TC', 'Match Rate']
    for col, h in enumerate(headers, 1):
        hdr_style(ws.cell(row=4, column=col, value=h))
    ws.row_dimensions[4].height = 24

    row_i = 5
    for label, res in results.items():
        if 'error' in res:
            vals = [label, '⚠️ ' + res['error'], '-', '-', '-', '-', '-']
            bg = 'FFE699'
        else:
            vals = [
                label,
                res['column_used'] or '-',
                res['total_mp_orders'],
                res['matched_count'],
                res.get('manual_confirmed_count', 0),
                res['missing_count'],
                f"{res['match_rate']}%" if res['match_rate'] is not None else '-',
            ]
            bg = GOOD_BG if res['missing_count'] == 0 else BAD_BG

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.font      = Font(name='Arial', bold=(col == 1), size=9)
            c.fill      = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='left' if col in (1, 2) else 'center', vertical='center')
            c.border    = border
        ws.row_dimensions[row_i].height = 20
        row_i += 1

    widths = [22, 22, 16, 14, 16, 16, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── SHEET 2+: MISSING ORDERS PER MARKETPLACE ─────────────────────────────
    for label, res in results.items():
        if 'error' in res or res['missing_count'] == 0:
            continue

        sheet_name = f"MISSING - {label}"[:31]  # Excel sheet name limit
        ws_m = wb.create_sheet(sheet_name)
        ws_m.sheet_view.showGridLines = False

        ws_m.merge_cells('A1:B1')
        ws_m['A1'].value     = f'{label} — Orders NOT found in TC ({res["missing_count"]} missing)'
        ws_m['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws_m['A1'].fill      = PatternFill('solid', start_color='C00000')
        ws_m['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_m.row_dimensions[1].height = 26

        hdr_style(ws_m.cell(row=2, column=1, value='#'), bg='C00000')
        hdr_style(ws_m.cell(row=2, column=2, value='Order ID / Order Number'), bg='C00000')
        ws_m.row_dimensions[2].height = 20

        for i, oid in enumerate(res['missing_ids'], 1):
            er = i + 2
            c1 = ws_m.cell(row=er, column=1, value=i)
            c2 = ws_m.cell(row=er, column=2, value=oid)
            for c in (c1, c2):
                c.font      = Font(name='Arial', size=9)
                c.fill      = PatternFill('solid', start_color=BAD_BG)
                c.alignment = Alignment(horizontal='center' if c == c1 else 'left', vertical='center')
                c.border    = border
            ws_m.row_dimensions[er].height = 18

        ws_m.column_dimensions['A'].width = 6
        ws_m.column_dimensions['B'].width = 30
        ws_m.freeze_panes = 'A3'

    # ── SHEET: MANUALLY CONFIRMED ─────────────────────────────────────────────
    if manual_df is not None and not manual_df.empty:
        ws_c = wb.create_sheet('MANUALLY CONFIRMED')
        ws_c.sheet_view.showGridLines = False

        cols = list(manual_df.columns)
        last_col_letter = get_column_letter(len(cols))

        ws_c.merge_cells(f'A1:{last_col_letter}1')
        ws_c['A1'].value     = f'Orders manually confirmed pushed to TC ({len(manual_df)} entries)'
        ws_c['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws_c['A1'].fill      = PatternFill('solid', start_color='375623')
        ws_c['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_c.row_dimensions[1].height = 26

        for col, h in enumerate(cols, 1):
            hdr_style(ws_c.cell(row=2, column=col, value=h), bg='375623')
        ws_c.row_dimensions[2].height = 22

        for i, (_, row) in enumerate(manual_df.iterrows(), 1):
            er = i + 2
            for col, h in enumerate(cols, 1):
                c = ws_c.cell(row=er, column=col, value=row[h])
                c.font      = Font(name='Arial', size=9)
                c.fill      = PatternFill('solid', start_color='E2EFDA')
                c.alignment = Alignment(horizontal='left', vertical='center')
                c.border    = border
            ws_c.row_dimensions[er].height = 18

        for i, h in enumerate(cols, 1):
            ws_c.column_dimensions[get_column_letter(i)].width = 20 if h != 'Marketplace' else 22
        ws_c.freeze_panes = 'A3'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
